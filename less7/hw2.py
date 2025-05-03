import os
import requests
import csv
from io import BytesIO
from zipfile import ZipFile
from selenium import webdriver
import PyPDF2
import pytest
from selene import browser
from selene import query
from selenium.webdriver.ie.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook



def test_text_in_downloaded_file():
    DIR_HW_ZIP = os.path.abspath('HW_zip')
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": DIR_HW_ZIP,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    browser.config.driver = driver

    browser.open('https://github.com/Persik78/Guru_HW/blob/master/less7/tmp/HW_zip%20final_final.zip')
    download_url_HW = browser.element('[data-testid="raw-button"]').get(query.attribute('href'))
    download_zip = requests.get(url=download_url_HW).content

    os.mkdir('HW_zip')

    with open('HW_zip/down_HW_zip.zip', 'wb') as zip_f:
        zip_f.write(download_zip)

    ZIP_DIR = os.path.join(DIR_HW_ZIP, 'down_HW_zip.zip')

    with ZipFile(ZIP_DIR) as zip_f:
        dir_list = zip_f.namelist()
        textInPDF = ''
        textInXLSX = ''
        textInCSV = ''
        for i in range(len(dir_list)):
            ext = os.path.splitext(dir_list[i])
            if ext[1] == '.pdf':
                with zip_f.open(dir_list[i]) as pdfFileInArachive:
                    content = BytesIO(pdfFileInArachive.read())
                    pdfReader = PyPDF2.PdfReader(content)

                    for page in pdfReader.pages:
                        textInPDF += page.extract_text() or ''

            elif ext[1] == '.xlsx':
                with zip_f.open(dir_list[i]) as xlsxFileInArchive:
                    content = BytesIO(xlsxFileInArchive.read())
                    xlsxReader = load_workbook(content)
                    sheet = xlsxReader.active
                    for row in sheet.iter_rows():
                        for cell in row:
                            textInXLSX += cell.value or ''

            elif ext[1] == '.csv':
                with zip_f.open(dir_list[i]) as csvFileInArchive:
                    content = csvFileInArchive.read().decode('utf-8-sig')
                    csvreader = list(csv.reader(content.splitlines()))
                    for row in csvreader:
                        textInCSV += str(row) or ''
    os.remove(ZIP_DIR)
    os.rmdir('HW_zip')

    assert 'TeSt_PdF_2_PAGE' in textInPDF
    assert 'TEST_PDF' in textInPDF
    assert 'asdasdasd' in textInXLSX
    assert 'adsasdada' in textInXLSX
    assert 'test_csv 0 0' in textInCSV
    assert 'test_csv 1 2' in textInCSV