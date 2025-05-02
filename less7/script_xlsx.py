from openpyxl import load_workbook

workbook = load_workbook('tmp/test_xlsx.xlsx')
sheet = workbook.active
print(sheet.cell(row=1, column=1).value)

for i in sheet.colu